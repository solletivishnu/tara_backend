import pandas as pd
from io import BytesIO
from datetime import date
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from rest_framework.permissions import AllowAny
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework.decorators import api_view, permission_classes
from .models import PayrollOrg, WorkLocations, Departments, Designation


@api_view(['GET'])
@permission_classes([AllowAny])
def generate_employee_upload_template(request, payroll_id):
    payroll = get_object_or_404(PayrollOrg, id=payroll_id)

    # Combined column headers from EmployeeManagement, EmployeePersonalDetails, EmployeeBankDetails
    columns = [
        # Basic employee
        'first_name', 'middle_name', 'last_name', 'associate_id', 'doj', 'work_email',
        'mobile_number', 'gender', 'work_location', 'designation', 'department',
        'enable_portal_access', 'epf_enabled', 'pf_account_number', 'uan',
        'esi_enabled', 'esi_number', 'professional_tax', 'employee_status',

        # Personal details
        'dob', 'age', 'guardian_name', 'pan', 'aadhar', 'address_line1',
        'address_line2', 'address_city', 'address_state', 'address_pinCode',
        'alternate_contact_number', 'marital_status', 'blood_group',

        # Bank details
        'account_holder_name', 'bank_name', 'account_number', 'ifsc_code', 'branch_name',
    ]

    # Create empty DataFrame
    df = pd.DataFrame(columns=columns)

    # Save to in-memory Excel file
    output = BytesIO()
    df.to_excel(output, index=False, sheet_name='EmployeeUpload')
    output.seek(0)

    # Load workbook and sheets
    wb = load_workbook(output)
    ws = wb['EmployeeUpload']
    dropdowns_ws = wb.create_sheet(title='Dropdowns')

    # Dropdown values
    genders = ['male', 'female', 'others']
    marital_status = ['single', 'married']
    blood_groups = ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-']
    boolean_values = ['TRUE', 'FALSE']
    work_locations = [str(wl) for wl in WorkLocations.objects.filter(payroll=payroll)]
    departments = [str(dept) for dept in Departments.objects.filter(payroll=payroll)]
    designations = [str(desig) for desig in Designation.objects.filter(payroll=payroll)]

    dropdowns_ws.append([
        'Genders', 'WorkLocations', 'Departments', 'Designations',
        'Boolean', 'MaritalStatus', 'BloodGroups'
    ])
    max_len = max(
        len(genders), len(work_locations), len(departments),
        len(designations), len(boolean_values), len(marital_status), len(blood_groups)
    )

    for i in range(max_len):
        dropdowns_ws.append([
            genders[i] if i < len(genders) else None,
            work_locations[i] if i < len(work_locations) else None,
            departments[i] if i < len(departments) else None,
            designations[i] if i < len(designations) else None,
            boolean_values[i] if i < len(boolean_values) else None,
            marital_status[i] if i < len(marital_status) else None,
            blood_groups[i] if i < len(blood_groups) else None,
        ])

    # Helper to apply dropdown
    def apply_dropdown(field_name, formula_range):
        for col in ws.iter_cols(1, ws.max_column):
            if col[0].value == field_name:
                col_letter = get_column_letter(col[0].column)
                dv = DataValidation(type="list", formula1=f"={formula_range}", allow_blank=False)
                ws.add_data_validation(dv)
                dv.add(f"{col_letter}2:{col_letter}1000")
                break

    # Helper to apply date validation
    def apply_date_validation(field_name):
        for col in ws.iter_cols(1, ws.max_column):
            if col[0].value == field_name:
                col_letter = get_column_letter(col[0].column)
                dv = DataValidation(type='date')
                dv.error = 'Please enter a valid date in YYYY-MM-DD format.'
                dv.errorTitle = 'Invalid Date'
                ws.add_data_validation(dv)
                dv.add(f'{col_letter}2:{col_letter}1000')
                ws.column_dimensions[col_letter].number_format = 'YYYY-MM-DD'
                break

    # Apply dropdowns
    apply_dropdown("gender", "Dropdowns!$A$2:$A$10")
    apply_dropdown("work_location", "Dropdowns!$B$2:$B$100")
    apply_dropdown("department", "Dropdowns!$C$2:$C$100")
    apply_dropdown("designation", "Dropdowns!$D$2:$D$100")
    apply_dropdown("enable_portal_access", "Dropdowns!$E$2:$E$3")
    apply_dropdown("epf_enabled", "Dropdowns!$E$2:$E$3")
    apply_dropdown("esi_enabled", "Dropdowns!$E$2:$E$3")
    apply_dropdown("professional_tax", "Dropdowns!$E$2:$E$3")
    apply_dropdown("employee_status", "Dropdowns!$E$2:$E$3")
    apply_dropdown("marital_status", "Dropdowns!$F$2:$F$3")
    apply_dropdown("blood_group", "Dropdowns!$G$2:$G$10")

    apply_date_validation('doj')
    apply_date_validation('dob')

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    # Finalize output
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    response = HttpResponse(
        final_output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="employee_upload_template.xlsx"'
    return response
