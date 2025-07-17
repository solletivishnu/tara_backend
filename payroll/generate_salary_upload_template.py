from io import BytesIO
from datetime import date
import pandas as pd
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.protection import SheetProtection
from .models import EmployeeManagement  # Update to your actual app name
from .helpers import default_earnings  # Your list of earnings
from rest_framework.permissions import AllowAny
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework.decorators import api_view, permission_classes
from .serializers import EmployeeSimpleSerializer, EarningsSerializer, DeductionSerializer
from openpyxl.styles import Font
from openpyxl.styles import Protection
from .models import *
from openpyxl import load_workbook
from django.db.models import Exists, OuterRef
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
# Sample default earnings (this can be moved to DB or settings)
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import date
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny

# Your default components and calculation logic
# default_earnings = [
#     {"component_name": "Basic", "calculation_type": {"type": "Percentage of CTC", "value": 50}},
#     {"component_name": "HRA", "calculation_type": {"type": "Percentage of Basic", "value": 50}},
#     {"component_name": "Fixed Allowance", "calculation_type": {"type": "Remaining balance pf CTC", "value": 0}},
#     {"component_name": "Conveyance Allowance", "calculation_type": {"type": "Flat Amount", "value": 0}},
#     {"component_name": "Bonus", "calculation_type": {"type": "Flat Amount", "value": 0}},
#     {"component_name": "Commission", "calculation_type": {"type": "Flat Amount", "value": 0}},
#     {"component_name": "Children Education Allowance", "calculation_type": {"type": "Flat Amount", "value": 0}},
#     {"component_name": "Transport Allowance", "calculation_type": {"type": "Flat Amount", "value": 0}},
#     {"component_name": "Travelling Allowance", "calculation_type": {"type": "Flat Amount", "value": 0}},
#     {"component_name": "Overtime Allowance", "calculation_type": {"type": "Flat Amount", "value": 0}},
# ]
#
# default_benefits = [
#     {"component_name": "EPF Employer Contribution", "calculation_type": {"type": "Percentage of PF wage", "value": 12}},
#     {"component_name": "EDLI Employer Contribution", "calculation_type": {"type": "Fixed Amount", "value": 75}},
#     {"component_name": "EPF Admin Charges", "calculation_type": {"type": "Fixed Amount", "value": 75}},
#     {"component_name": "ESI Employer Contribution", "calculation_type": {"type": "Not Applicable", "value": None}},
# ]
#
# default_deductions = [
#     {"component_name": "EPF Employee Contribution", "calculation_type": {"type": "Percentage of PF wage", "value": 12}},
#     {"component_name": "ESI Employee Contribution", "calculation_type": {"type": "Not Applicable", "value": None}},
#     {"component_name": "Professional Tax (PT)", "calculation_type": {"type": "Not Applicable", "value": None}},
# ]
#
#
# @api_view(['GET'])
# @permission_classes([AllowAny])
# def generate_salary_upload_template(request, payroll_id):
#     try:
#         payroll = PayrollOrg.objects.get(pk=payroll_id)
#         # Exclude employees who already have a salary record
#         employees = EmployeeManagement.objects.filter(
#             payroll=payroll
#         ).annotate(
#             has_salary=Exists(EmployeeSalaryDetails.objects.filter(employee=OuterRef('pk')))
#         ).filter(has_salary=False).order_by('id')
#         serializer = EmployeeSimpleSerializer(employees, many=True)
#
#         base_columns = ['employee_id', 'associate_id', 'full_name', 'annual_ctc', 'tax_regime_opted', 'valid_from']
#
#         earning_fields = ['component_name', 'calculation', 'calculation_type', 'monthly', 'annually']
#         earning_columns = [f'earning_{i}_{field}' for i in range(1, len(default_earnings) + 1)
#                            for field in earning_fields]
#
#         benefit_fields = ['component_name', 'calculation', 'calculation_type', 'monthly', 'annually']
#         benefit_columns = [f'benefit_{i}_{field}' for i in range(1, len(default_benefits) + 1)
#                            for field in benefit_fields]
#
#         deduction_fields = ['component_name', 'calculation', 'calculation_type', 'monthly', 'annually']
#         deduction_columns = [f'deduction_{i}_{field}' for i in range(1, len(default_deductions) + 1)
#                              for field in deduction_fields]
#
#         summary_columns = [
#             'gross_salary_monthly', 'gross_salary_annually',
#             'total_ctc_monthly', 'total_ctc_annually',
#             'net_salary_monthly', 'net_salary_annually'
#         ]
#
#         final_columns = base_columns + earning_columns + benefit_columns + deduction_columns + summary_columns
#
#         data = []
#         for emp in serializer.data:
#             row = {
#                 'employee_id': emp['id'],
#                 'associate_id': emp['associate_id'],
#                 'full_name': emp['full_name'],
#                 'annual_ctc': '',
#                 'tax_regime_opted': '',
#                 'valid_from': ''
#             }
#
#             for i, earning in enumerate(default_earnings, start=1):
#                 row[f'earning_{i}_component_name'] = earning['component_name']
#                 row[f'earning_{i}_calculation_type'] = earning['calculation_type']['type']
#                 row[f'earning_{i}_calculation'] = '' \
#                     if earning['calculation_type']['type'] in ["Flat Amount", "Remaining balance pf CTC"] \
#                     else earning['calculation_type']['value']
#                 row[f'earning_{i}_monthly'] = ''
#                 row[f'earning_{i}_annually'] = ''
#
#             for i, benefit in enumerate(default_benefits, start=1):
#                 row[f'benefit_{i}_component_name'] = benefit['component_name']
#                 row[f'benefit_{i}_calculation_type'] = benefit['calculation_type']['type']
#                 row[f'benefit_{i}_calculation'] = '' \
#                     if benefit['calculation_type']['type'] in ["Flat Amount", "Not Applicable"] \
#                     else benefit['calculation_type']['value']
#                 row[f'benefit_{i}_monthly'] = ''
#                 row[f'benefit_{i}_annually'] = ''
#
#             for i, deduction in enumerate(default_deductions, start=1):
#                 row[f'deduction_{i}_component_name'] = deduction['component_name']
#                 row[f'deduction_{i}_calculation_type'] = deduction['calculation_type']['type']
#                 row[f'deduction_{i}_calculation'] = '' \
#                     if deduction['calculation_type']['type'] in ["Flat Amount", "Not Applicable"] \
#                     else deduction['calculation_type']['value']
#                 row[f'deduction_{i}_monthly'] = ''
#                 row[f'deduction_{i}_annually'] = ''
#
#             for col in summary_columns:
#                 row[col] = ''
#
#             data.append(row)
#
#         df = pd.DataFrame(data, columns=final_columns)
#
#         output = BytesIO()
#         with pd.ExcelWriter(output, engine='openpyxl') as writer:
#             df.to_excel(writer, index=False, sheet_name='SalaryTemplate')
#
#         output.seek(0)
#         wb = load_workbook(output)
#         ws = wb.active
#         ws.freeze_panes = "A2"
#
#         for col in ws.columns:
#             max_length = 0
#             col_letter = col[0].column_letter
#             for cell in col:
#                 if cell.value:
#                     max_length = max(max_length, len(str(cell.value)))
#             ws.column_dimensions[col_letter].width = max_length + 2
#
#         def col_index(name):
#             return final_columns.index(name) + 1
#
#         n_rows = ws.max_row
#
#         editable_columns = ['annual_ctc', 'tax_regime_opted', 'valid_from']
#         editable_columns += [f'earning_{i}_calculation' for i, e in enumerate(default_earnings, 1)
#                              if e['calculation_type']['type'] in ['Flat Amount', 'Remaining balance pf CTC',
#                                                                   'Percentage of CTC', 'Percentage of Basic']]
#         editable_columns += [f'earning_{i}_monthly' for i in range(1, len(default_earnings)+1)]
#         editable_columns += [f'earning_{i}_annually' for i in range(1, len(default_earnings)+1)]
#         editable_columns += [f'benefit_{i}_calculation' for i, b in enumerate(default_benefits, 1)
#                              if b['calculation_type']['type'] in ['Flat Amount']]
#         editable_columns += [f'benefit_{i}_monthly' for i in range(1, len(default_benefits)+1)]
#         editable_columns += [f'benefit_{i}_annually' for i in range(1, len(default_benefits)+1)]
#         editable_columns += [f'deduction_{i}_calculation' for i, d in enumerate(default_deductions, 1)
#                              if d['calculation_type']['type'] in ['Flat Amount']]
#         editable_columns += [f'deduction_{i}_monthly' for i in range(1, len(default_deductions)+1)]
#         editable_columns += [f'deduction_{i}_annually' for i in range(1, len(default_deductions)+1)]
#
#         for col_idx, col_name in enumerate(final_columns, start=1):
#             for row_idx in range(2, n_rows+1):
#                 cell = ws.cell(row=row_idx, column=col_idx)
#                 cell.protection = cell.protection.copy(locked=col_name not in editable_columns)
#
#         for row_idx in range(2, n_rows + 1):
#             annual_ctc_cell = ws.cell(row=row_idx, column=col_index('annual_ctc')).coordinate
#
#             # EARNINGS FORMULAS (dynamic for all earnings)
#             # Remove hardcoded Basic, HRA, and 'other earnings' loop
#             for earning in default_earnings:
#                 ct = earning['calculation_type']['type']
#                 comp = earning['component_name']
#                 # Find input and output columns
#                 if ct == "Percentage of CTC":
#                     input_col = col_index(f'{comp} (% of CTC)')
#                 elif ct == "Percentage of Basic":
#                     input_col = col_index(f'{comp} (% of Basic)')
#                 elif ct == "Flat Amount":
#                     input_col = col_index(f'{comp} (Flat Amount)')
#                 else:
#                     input_col = None
#                 monthly_col = col_index(f'Monthly ({comp})')
#                 annually_col = col_index(f'Annually ({comp})')

#                 # Assign formulas
#                 if ct == "Percentage of CTC" and input_col:
#                     ws.cell(row=row_idx, column=monthly_col).value = (
#                         f"=IF(AND(ISNUMBER({annual_ctc_cell}), ISNUMBER({get_column_letter(input_col)}{row_idx})), "
#                         f"{annual_ctc_cell}*{get_column_letter(input_col)}{row_idx}/100/12, \"\")"
#                     )
#                 elif ct == "Percentage of Basic" and input_col:
#                     # Find the basic monthly column
#                     basic_monthly_col = None
#                     for e2 in default_earnings:
#                         if e2['component_name'].lower() == 'basic':
#                             basic_monthly_col = col_index(f'Monthly ({e2["component_name"]})')
#                             break
#                     if basic_monthly_col:
#                         ws.cell(row=row_idx, column=monthly_col).value = (
#                             f"=IF(AND(ISNUMBER({get_column_letter(basic_monthly_col)}{row_idx}), ISNUMBER({get_column_letter(input_col)}{row_idx})), "
#                             f"{get_column_letter(basic_monthly_col)}{row_idx}*{get_column_letter(input_col)}{row_idx}/100, \"\")"
#                         )
#                 elif ct == "Flat Amount" and input_col:
#                     ws.cell(row=row_idx, column=monthly_col).value = (
#                         f"=IF(ISNUMBER({get_column_letter(input_col)}{row_idx}), {get_column_letter(input_col)}{row_idx}, \"\")"
#                     )
#                 elif ct == "Remaining balance pf CTC":
#                     # Calculate sum of all other earnings and benefits
#                     other_earnings_sum = []
#                     for e2 in default_earnings:
#                         if e2['component_name'] != comp:
#                             monthly_col2 = col_index(f'Monthly ({e2["component_name"]})')
#                             other_earnings_sum.append(
#                                 f"IF(ISNUMBER({get_column_letter(monthly_col2)}{row_idx}), {get_column_letter(monthly_col2)}{row_idx}, 0)")
#                     other_benefits_sum = []
#                     for benefit in default_benefits:
#                         monthly_col2 = col_index(f'Monthly ({benefit["component_name"]})')
#                         other_benefits_sum.append(
#                             f"IF(ISNUMBER({get_column_letter(monthly_col2)}{row_idx}), {get_column_letter(monthly_col2)}{row_idx}, 0)")
#                     total_other = " + ".join(other_earnings_sum + other_benefits_sum) if (other_earnings_sum + other_benefits_sum) else "0"
#                     ws.cell(row=row_idx, column=monthly_col).value = (
#                         f"=IF(ISNUMBER({annual_ctc_cell}), "
#                         f"IF(({annual_ctc_cell}/12) - ({total_other}) >= 0, "
#                         f"({annual_ctc_cell}/12) - ({total_other}), "
#                         f"\"Error: Adjust earnings\"), \"\")"
#                     )
#                 # Annually = Monthly * 12 for all
#                 ws.cell(row=row_idx, column=annually_col).value = (
#                     f"=IF(ISNUMBER({get_column_letter(monthly_col)}{row_idx}), {get_column_letter(monthly_col)}{row_idx}*12, \"\")"
#                 )

#             # Benefits
#             for i, benefit in enumerate(default_benefits, start=1):
#                 ct = benefit['calculation_type']['type']
#                 val = benefit['calculation_type']['value']
#                 m_col = col_index(f'benefit_{i}_monthly')
#                 a_col = col_index(f'benefit_{i}_annually')
#                 m_cell = ws.cell(row=row_idx, column=m_col)
#                 a_cell = ws.cell(row=row_idx, column=a_col)
#                 if ct == "Percentage of PF wage":
#                     basic = ws.cell(row=row_idx, column=col_index('earning_1_monthly')).coordinate
#                     m_cell.value = f"=IF(ISNUMBER({basic}), MIN({basic},15000)*{val}/100,\"\")"
#                     a_cell.value = f"=IF(ISNUMBER({m_cell.coordinate}), {m_cell.coordinate}*12,\"\")"
#                 elif ct == "Fixed Amount":
#                     c_cell = ws.cell(row=row_idx, column=col_index(f'benefit_{i}_calculation')).coordinate
#                     m_cell.value = f"={c_cell}"
#                     a_cell.value = f"=IF(ISNUMBER({m_cell.coordinate}), {m_cell.coordinate}*12,\"\")"

#             # Deductions
#             for i, deduction in enumerate(default_deductions, start=1):
#                 ct = deduction['calculation_type']['type']
#                 val = deduction['calculation_type']['value']
#                 component = deduction.get('component_name', '').strip().lower()

#                 # Skip if component is missing or calculation type is 'Not Applicable'
#                 if not component or ct.lower() == "not applicable":
#                     continue

#                 m_col = col_index(f'deduction_{i}_monthly')
#                 a_col = col_index(f'deduction_{i}_annually')
#                 m_cell = ws.cell(row=row_idx, column=m_col)
#                 a_cell = ws.cell(row=row_idx, column=a_col)

#                 basic_cell = ws.cell(row=row_idx, column=col_index('earning_1_monthly')).coordinate
#                 pf_wage_expr = f"MIN({basic_cell},15000)"

#                 # Component-based logic
#                 if component == "esi employee contribution":
#                     m_cell.value = (
#                         f'=IF(AND(ISNUMBER({basic_cell}), {basic_cell}<=21000), '
#                         f'ROUND({pf_wage_expr}*0.0075, 2), "")'
#                     )
#                     a_cell.value = f'=IF(ISNUMBER({m_cell.coordinate}), ROUND({m_cell.coordinate}*12, 2), "")'

#                 elif component == "pt":
#                     m_cell.value = '200'
#                     a_cell.value = f'=IF(ISNUMBER({m_cell.coordinate}), {m_cell.coordinate}*12, "")'

#                 elif ct == "Percentage of PF wage":
#                     m_cell.value = f'=IF(ISNUMBER({basic_cell}), {pf_wage_expr}*{val}/100, "")'
#                     a_cell.value = f'=IF(ISNUMBER({m_cell.coordinate}), {m_cell.coordinate}*12, "")'

#                 elif ct == "Fixed Amount":
#                     m_cell.value = f'={val}'
#                     a_cell.value = f'=IF(ISNUMBER({m_cell.coordinate}), {m_cell.coordinate}*12, "")'

#                 else:
#                     # Fallback (if needed)
#                     m_cell.value = ""
#                     a_cell.value = ""

#             # Summary
#             earn_start = col_index('earning_1_monthly')
#             earn_end = col_index(f'earning_{len(default_earnings)}_monthly')
#             gross_m = col_index('gross_salary_monthly')
#             gross_a = col_index('gross_salary_annually')

#             benefit_start = col_index('benefit_1_monthly')
#             benefit_end = col_index(f'benefit_{len(default_benefits)}_monthly')
#             total_ctc_m = col_index('total_ctc_monthly')
#             total_ctc_a = col_index('total_ctc_annually')

#             deduct_start = col_index('deduction_1_monthly')
#             deduct_end = col_index(f'deduction_{len(default_deductions)}_monthly')
#             net_m = col_index('net_salary_monthly')
#             net_a = col_index('net_salary_annually')

#             annual_ctc_cell = ws.cell(row=row_idx, column=col_index('annual_ctc')).coordinate

#             ws.cell(row=row_idx, column=col_index('gross_salary_monthly')).value = \
#                 (f"=IF(ISNUMBER({annual_ctc_cell}), "
#                  f"SUM(J{row_idx}, O{row_idx}, T{row_idx}, Y{row_idx}, "
#                  f"AD{row_idx}, AI{row_idx}, AN{row_idx}, AS{row_idx}, "
#                  f"AX{row_idx}, BC{row_idx}), \"\")")

#             ws.cell(row=row_idx, column=col_index('gross_salary_annually')).value = \
#                 (f"=IF(ISNUMBER({annual_ctc_cell}), {get_column_letter(col_index('gross_salary_monthly'))}"
#                  f"{row_idx}*12, \"\")")

#             ws.cell(row=row_idx, column=col_index('total_ctc_monthly')).value = \
#                 (f"=IF(ISNUMBER({annual_ctc_cell}), "
#                  f"{get_column_letter(col_index('gross_salary_monthly'))}{row_idx}"
#                  f"+SUM(BH{row_idx}, BM{row_idx}, BR{row_idx}, BW{row_idx}), \"\")")

#             ws.cell(row=row_idx, column=col_index('total_ctc_annually')).value = \
#                 (f"=IF(ISNUMBER({annual_ctc_cell}), "
#                  f"{get_column_letter(col_index('total_ctc_monthly'))}{row_idx}*12, \"\")")

#             ws.cell(row=row_idx, column=col_index('net_salary_monthly')).value = \
#                 (f"=IF(ISNUMBER({annual_ctc_cell}), "
#                  f"{get_column_letter(col_index('total_ctc_monthly'))}"
#                  f"{row_idx}-SUM(CB{row_idx}, CG{row_idx}, CL{row_idx}), \"\")")

#             ws.cell(row=row_idx, column=col_index('net_salary_annually')).value = \
#                 (f"=IF(ISNUMBER({annual_ctc_cell}), "
#                  f"{get_column_letter(col_index('net_salary_monthly'))}{row_idx}*12, \"\")")

#         # Find the column index for 'tax_regime_opted'
#         tax_regime_col_idx = final_columns.index('tax_regime_opted') + 1  # 1-based index
#         n_rows = ws.max_row

#         # Create the data validation dropdown
#         dv = DataValidation(type="list", formula1='"old,new"', allow_blank=True)
#         dv.error = 'Select either "old" or "new"'
#         dv.errorTitle = 'Invalid Input'

#         # The range for the dropdown (excluding header)
#         col_letter = get_column_letter(tax_regime_col_idx)
#         dv_range = f"{col_letter}2:{col_letter}{n_rows}"

#         dv.add(dv_range)
#         ws.add_data_validation(dv)

#         ws.protection.sheet = True
#         ws.protection.password = 'tara'
#         ws.protection.enable()

#         buffer = BytesIO()
#         wb.save(buffer)
#         buffer.seek(0)

#         filename = f"EmployeeSalaryTemplate_{date.today().strftime('%Y-%m-%d')}.xlsx"
#         return HttpResponse(buffer.read(),
#                             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#                             headers={'Content-Disposition': f'attachment; filename="{filename}"'})

#     except PayrollOrg.DoesNotExist:
#         return HttpResponse("Payroll not found", status=404)


default_earnings = [
    {"component_name": "Basic", "calculation_type": {"type": "Percentage of CTC", "value": 50}},
    {"component_name": "HRA", "calculation_type": {"type": "Percentage of Basic", "value": 50}},
    {"component_name": "Fixed Allowance", "calculation_type": {"type": "Remaining balance pf CTC", "value": 0}},
    {"component_name": "Conveyance Allowance", "calculation_type": {"type": "Flat Amount", "value": 0}},
    {"component_name": "Bonus", "calculation_type": {"type": "Flat Amount", "value": 0}},
    {"component_name": "Commission", "calculation_type": {"type": "Flat Amount", "value": 0}},
    {"component_name": "Children Education Allowance", "calculation_type": {"type": "Flat Amount", "value": 0}},
    {"component_name": "Transport Allowance", "calculation_type": {"type": "Flat Amount", "value": 0}},
    {"component_name": "Travelling Allowance", "calculation_type": {"type": "Flat Amount", "value": 0}},
    {"component_name": "Overtime Allowance", "calculation_type": {"type": "Flat Amount", "value": 0}},
]

default_benefits = [
    {"component_name": "EPF Employer Contribution", "calculation_type": {"type": "Percentage of PF wage", "value": 12}},
    {"component_name": "EDLI Employer Contribution", "calculation_type": {"type": "Fixed Amount", "value": 75}},
    {"component_name": "EPF Admin Charges", "calculation_type": {"type": "Fixed Amount", "value": 75}},
    {"component_name": "ESI Employer Contribution", "calculation_type": {"type": "Percentage of PF wage", "value": 3.25}
     },
]

default_deductions = [
    {"component_name": "EPF Employee Contribution", "calculation_type": {"type": "Percentage of PF wage", "value": 12}},
    {"component_name": "ESI Employee Contribution", "calculation_type": {"type": "Percentage of PF wage", "value": 0.75}
     },
    {"component_name": "Professional Tax (PT)", "calculation_type":
        {"type": "Based On Monthly Basic Criteria", "value": 0}},
]


@api_view(['GET'])
@permission_classes([AllowAny])
def generate_salary_upload_template(request, payroll_id):
    try:
        payroll = PayrollOrg.objects.get(pk=payroll_id)
        employees = EmployeeManagement.objects.filter(
            payroll=payroll
        ).annotate(
            has_salary=Exists(EmployeeSalaryDetails.objects.filter(employee=OuterRef('pk')))
        ).filter(has_salary=False).order_by('id')
        serializer = EmployeeSimpleSerializer(employees, many=True)

        if not employees.exists():
            return HttpResponse("No employees found without salary details.", status=204)

        # New column structure
        base_columns = ['Employee Id', 'Associate Id', 'Full name', 'Annual CTC', 'Tax Regime Opted']

        # Earnings columns - each has input column + monthly + annually
        earning_columns = []
        earnings = EarningsSerializer(Earnings.objects.filter(payroll=payroll).order_by('id'), many=True).data
        for earning in earnings:
            ct = earning['calculation_type']['type']
            if ct == "Percentage of CTC":
                earning_columns.append(f"{earning['component_name']} (% of CTC)")
            elif ct == "Percentage of Basic":
                earning_columns.append(f"{earning['component_name']} (% of Basic)")
            elif ct == "Flat Amount":
                earning_columns.append(f"{earning['component_name']} (Flat Amount)")

            earning_columns.append(f"Monthly ({earning['component_name']})")
            earning_columns.append(f"Annually ({earning['component_name']})")

        # Benefits columns - only monthly and annually
        benefit_columns = []
        for benefit in default_benefits:
            benefit_columns.append(f"Monthly ({benefit['component_name']})")
            benefit_columns.append(f"Annually ({benefit['component_name']})")

        # Deductions columns - only monthly and annually
        deduction_columns = []
        for deduction in default_deductions:
            deduction_columns.append(f"Monthly ({deduction['component_name']})")
            deduction_columns.append(f"Annually ({deduction['component_name']})")

        deduction_data = DeductionSerializer(Deduction.objects.filter(payroll=payroll).order_by('id'), many=True).data
        for deduction_row in deduction_data:
            ct = deduction_row['calculation_type']['type']
            if ct == "Percentage of CTC":
                deduction_columns.append(f"{deduction_row['deduction_name']} (% of CTC)")
            elif ct == "Percentage of Basic":
                deduction_columns.append(f"{deduction_row['deduction_name']} (% of Basic)")
            elif ct == "Flat Amount":
                deduction_columns.append(f"{deduction_row['deduction_name']} (Flat Amount)")

            deduction_columns.append(f"Monthly ({deduction_row['deduction_name']})")
            deduction_columns.append(f"Annually ({deduction_row['deduction_name']})")

        summary_columns = [
            'Gross Salary (Monthly)', 'Gross Salary (Annually)',
            'Total CTC (Monthly)', 'Total CTC (Annually)',
            'Net Salary (Monthly)', 'Net Salary (Annually)'
        ]

        final_columns = base_columns + earning_columns + benefit_columns + deduction_columns + summary_columns

        data = []
        for emp in serializer.data:
            row = {
                'Employee Id': emp['id'],
                'Associate Id': emp['associate_id'],
                'Full name': emp['full_name'],
                'Annual CTC': '',
                'Tax Regime Opted': '',
            }

            # Initialize all columns with empty values
            for col in final_columns[len(base_columns):]:
                row[col] = ''

            # Set default values for specific columns
            for i, earning in enumerate(earnings, start=1):
                ct = earning['calculation_type']['type']
                value = earning['calculation_type'].get('value', '')
                # Convert value to number if it's a string and represents a number
                if isinstance(value, str):
                    try:
                        if '.' in value:
                            value = float(value)
                        else:
                            value = int(value)
                    except Exception:
                        pass  # leave as string if not convertible
                if ct == "Percentage of CTC":
                    row[f"{earning['component_name']} (% of CTC)"] = value
                elif ct == "Percentage of Basic":
                    row[f"{earning['component_name']} (% of Basic)"] = value

            data.append(row)

        df = pd.DataFrame(data, columns=final_columns)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='SalaryTemplate')

        output.seek(0)
        wb = load_workbook(output)
        ws = wb.active
        ws.freeze_panes = "A2"

        header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

        # Only apply styling if at least one row (i.e., headers) exists
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.fill = header_fill

        # Set column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        def col_index(name):
            return final_columns.index(name) + 1  # +1 because Excel is 1-based

        n_rows = ws.max_row

        # Define editable columns
        editable_columns = ['Annual CTC', 'Tax Regime Opted']

        # Add earning input columns
        for earning in earnings:
            ct = earning['calculation_type']['type']
            if ct == "Percentage of CTC":
                editable_columns.append(f"{earning['component_name']} (% of CTC)")
            elif ct == "Percentage of Basic":
                editable_columns.append(f"{earning['component_name']} (% of Basic)")
            elif ct == "Flat Amount":
                editable_columns.append(f"{earning['component_name']} (Flat Amount)")
        for deduction_row in deduction_data:
            ct = deduction_row['calculation_type']['type']
            if ct == "Flat Amount":
                editable_columns.append(f"{deduction_row['deduction_name']} (Flat Amount)")

        # Set cell protection
        for col_idx, col_name in enumerate(final_columns, start=1):
            for row_idx in range(2, n_rows + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.protection = cell.protection.copy(locked=col_name not in editable_columns)

        # Add formulas for each employee row
        for row_idx in range(2, n_rows + 1):
            annual_ctc_cell = ws.cell(row=row_idx, column=col_index('Annual CTC')).coordinate

            # EARNINGS FORMULAS (dynamic for all earnings)
            # Remove hardcoded Basic, HRA, and 'other earnings' loop
            for earning in earnings:
                ct = earning['calculation_type']['type']
                comp = earning['component_name']
                # Find input and output columns
                if ct == "Percentage of CTC":
                    input_col = col_index(f'{comp} (% of CTC)')
                elif ct == "Percentage of Basic":
                    input_col = col_index(f'{comp} (% of Basic)')
                elif ct == "Flat Amount":
                    input_col = col_index(f'{comp} (Flat Amount)')
                else:
                    input_col = None
                monthly_col = col_index(f'Monthly ({comp})')
                annually_col = col_index(f'Annually ({comp})')

                # Assign formulas
                if ct == "Percentage of CTC" and input_col:
                    ws.cell(row=row_idx, column=monthly_col).value = (
                        f"=IF(AND(ISNUMBER({annual_ctc_cell}), ISNUMBER({get_column_letter(input_col)}{row_idx})), "
                        f"{annual_ctc_cell}*{get_column_letter(input_col)}{row_idx}/100/12, \"\")"
                    )
                elif ct == "Percentage of Basic" and input_col:
                    # Find the basic monthly column
                    basic_monthly_col = None
                    for e2 in earnings:
                        if e2['component_name'].lower() == 'basic':
                            basic_monthly_col = col_index(f'Monthly ({e2["component_name"]})')
                            break
                    if basic_monthly_col:
                        ws.cell(row=row_idx, column=monthly_col).value = (
                            f"=IF(AND(ISNUMBER({get_column_letter(basic_monthly_col)}{row_idx}), ISNUMBER({get_column_letter(input_col)}{row_idx})), "
                            f"{get_column_letter(basic_monthly_col)}{row_idx}*{get_column_letter(input_col)}{row_idx}/100, \"\")"
                        )
                elif ct == "Flat Amount" and input_col:
                    ws.cell(row=row_idx, column=monthly_col).value = (
                        f"=IF(ISNUMBER({get_column_letter(input_col)}{row_idx}), {get_column_letter(input_col)}{row_idx}, \"\")"
                    )
                elif ct == "Remaining balance pf CTC":
                    # Calculate sum of all other earnings and benefits
                    other_earnings_sum = []
                    for e2 in earnings:
                        if e2['component_name'] != comp:
                            monthly_col2 = col_index(f'Monthly ({e2["component_name"]})')
                            other_earnings_sum.append(
                                f"IF(ISNUMBER({get_column_letter(monthly_col2)}{row_idx}), {get_column_letter(monthly_col2)}{row_idx}, 0)")
                    other_benefits_sum = []
                    for benefit in default_benefits:
                        monthly_col2 = col_index(f'Monthly ({benefit["component_name"]})')
                        other_benefits_sum.append(
                            f"IF(ISNUMBER({get_column_letter(monthly_col2)}{row_idx}), {get_column_letter(monthly_col2)}{row_idx}, 0)")
                    total_other = " + ".join(other_earnings_sum + other_benefits_sum) if (other_earnings_sum + other_benefits_sum) else "0"
                    ws.cell(row=row_idx, column=monthly_col).value = (
                        f"=IF(ISNUMBER({annual_ctc_cell}), "
                        f"IF(ABS(({annual_ctc_cell}/12) - ({total_other})) < 0.01, 0, "
                        f"IF(({annual_ctc_cell}/12) - ({total_other}) >= 0, "
                        f"({annual_ctc_cell}/12) - ({total_other}), "
                        f"\"Error: Adjust earnings\")), \"\")"
                    )

                # Annually = Monthly * 12 for all
                ws.cell(row=row_idx, column=annually_col).value = (
                    f"=IF(ISNUMBER({get_column_letter(monthly_col)}{row_idx}), {get_column_letter(monthly_col)}{row_idx}*12, \"\")"
                )

            # BENEFITS FORMULAS
            # 1. EPF Employer Contribution (12% of PF wage)
            epf_employer_monthly_col = col_index('Monthly (EPF Employer Contribution)')
            ws.cell(row=row_idx, column=epf_employer_monthly_col).value = (
                f"=IF(ISNUMBER({get_column_letter(col_index('Monthly (Basic)'))}{row_idx}), "
                f"MIN({get_column_letter(col_index('Monthly (Basic)'))}{row_idx},15000)*0.12, \"\")"
            )
            ws.cell(row=row_idx, column=col_index('Annually (EPF Employer Contribution)')).value = (
                f"=IF(ISNUMBER({get_column_letter(epf_employer_monthly_col)}{row_idx}), "
                f"{get_column_letter(epf_employer_monthly_col)}{row_idx}*12, \"\")"
            )

            # 2. EDLI and EPF Admin Charges (Fixed Amounts)
            for benefit in default_benefits[1:3]:  # EDLI and EPF Admin
                monthly_col = col_index(f'Monthly ({benefit["component_name"]})')
                ws.cell(row=row_idx, column=monthly_col).value = (
                    f"=IF(ISNUMBER({annual_ctc_cell}), {benefit['calculation_type']['value']}, \"\")"
                )
                ws.cell(row=row_idx, column=col_index(f'Annually ({benefit["component_name"]})')).value = (
                    f"=IF(ISNUMBER({get_column_letter(monthly_col)}{row_idx}), "
                    f"{get_column_letter(monthly_col)}{row_idx}*12, \"\")"
                )

            # 3. Deductions Formulas
            for deduction_row in deduction_data:
                monthly_col = col_index(f'Monthly ({deduction_row["deduction_name"]})')
                annually_col = col_index(f'Annually ({deduction_row["deduction_name"]})')
                comp = deduction_row['deduction_name']
                ct = deduction_row['calculation_type']['type']
                if ct == "Flat Amount":
                    input_col = col_index(f'{comp} (Flat Amount)')
                if ct == "Flat Amount" and input_col:
                    ws.cell(row=row_idx, column=monthly_col).value = (
                        f"=IF(ISNUMBER({get_column_letter(input_col)}{row_idx}), {get_column_letter(input_col)}{row_idx}, \"\")"
                    )
                # Annually = Monthly * 12 for all
                ws.cell(row=row_idx, column=annually_col).value = (
                    f"=IF(ISNUMBER({get_column_letter(monthly_col)}{row_idx}), {get_column_letter(monthly_col)}{row_idx}*12, \"\")"
                )

            # DEDUCTIONS FORMULAS
            # 1. EPF Employee Contribution (12% of PF wage)
            epf_employee_monthly_col = col_index('Monthly (EPF Employee Contribution)')
            ws.cell(row=row_idx, column=epf_employee_monthly_col).value = (
                f"=IF(ISNUMBER({get_column_letter(col_index('Monthly (Basic)'))}{row_idx}), "
                f"MIN({get_column_letter(col_index('Monthly (Basic)'))}{row_idx},15000)*0.12, \"\")"
            )
            ws.cell(row=row_idx, column=col_index('Annually (EPF Employee Contribution)')).value = (
                f"=IF(ISNUMBER({get_column_letter(epf_employee_monthly_col)}{row_idx}), "
                f"{get_column_letter(epf_employee_monthly_col)}{row_idx}*12, \"\")"
            )

            # 2. Professional Tax (Fixed Amount)
            pt_monthly_col = col_index('Monthly (Professional Tax (PT))')
            ws.cell(row=row_idx, column=pt_monthly_col).value = (
                f"=IF(ISNUMBER({annual_ctc_cell}), 200, \"\")"
            )
            ws.cell(row=row_idx, column=col_index('Annually (Professional Tax (PT))')).value = (
                f"=IF(ISNUMBER({get_column_letter(pt_monthly_col)}{row_idx}), "
                f"{get_column_letter(pt_monthly_col)}{row_idx}*12, \"\")"
            )

            esi_employer_monthly_col = col_index('Monthly (ESI Employer Contribution)')

            ws.cell(row=row_idx, column=esi_employer_monthly_col).value = (
                f"=IF(ISNUMBER({annual_ctc_cell}), "
                f"IF(AND(ISNUMBER({get_column_letter(col_index('Monthly (Basic)'))}{row_idx}), "
                f"{get_column_letter(col_index('Monthly (Basic)'))}{row_idx}<=21000), "
                f"ROUND(MIN({get_column_letter(col_index('Monthly (Basic)'))}{row_idx},15000)*0.0325, 2), \"\"), \"\")"
            )

            ws.cell(row=row_idx, column=col_index('Annually (ESI Employer Contribution)')).value = (
                f"=IF(ISNUMBER({get_column_letter(esi_employer_monthly_col)}{row_idx}), "
                f"ROUND({get_column_letter(esi_employer_monthly_col)}{row_idx}*12, 2), \"\")"
            )

            # ESI Employee Contribution (0.75% when Basic <= 21000)
            esi_employee_monthly_col = col_index('Monthly (ESI Employee Contribution)')

            ws.cell(row=row_idx, column=esi_employee_monthly_col).value = (
                f"=IF(ISNUMBER({annual_ctc_cell}), "
                f"IF(AND(ISNUMBER({get_column_letter(col_index('Monthly (Basic)'))}{row_idx}), "
                f"{get_column_letter(col_index('Monthly (Basic)'))}{row_idx}<=21000), "
                f"ROUND(MIN({get_column_letter(col_index('Monthly (Basic)'))}{row_idx},15000)*0.0075, 2), \"\"), \"\")"
            )

            ws.cell(row=row_idx, column=col_index('Annually (ESI Employee Contribution)')).value = (
                f"=IF(ISNUMBER({get_column_letter(esi_employee_monthly_col)}{row_idx}), "
                f"ROUND({get_column_letter(esi_employee_monthly_col)}{row_idx}*12, 2), \"\")"
            )

            # SUMMARY FORMULAS
            # Gross Salary (sum of all earnings)
            earnings_sum = []
            for earning in earnings:
                monthly_col = col_index(f'Monthly ({earning["component_name"]})')
                earnings_sum.append(
                    f"IF(ISNUMBER({get_column_letter(monthly_col)}{row_idx}), {get_column_letter(monthly_col)}{row_idx}, 0)")

            ws.cell(row=row_idx, column=col_index('Gross Salary (Monthly)')).value = (
                    f"=IF(ISNUMBER({annual_ctc_cell}), SUM({', '.join(earnings_sum)}), \"\")"
                )
            ws.cell(row=row_idx, column=col_index('Gross Salary (Annually)')).value = (
                f"=IF(ISNUMBER({get_column_letter(col_index('Gross Salary (Monthly)'))}{row_idx}), "
                f"{get_column_letter(col_index('Gross Salary (Monthly)'))}{row_idx}*12, \"\")"
            )

            # Total CTC (Gross Salary + Benefits)
            benefits_sum = []
            for benefit in default_benefits:
                monthly_col = col_index(f'Monthly ({benefit["component_name"]})')
                benefits_sum.append(
                    f"IF(ISNUMBER({get_column_letter(monthly_col)}{row_idx}), {get_column_letter(monthly_col)}{row_idx}, 0)")

            ws.cell(row=row_idx, column=col_index('Total CTC (Monthly)')).value = (
                f"=IF(ISNUMBER({annual_ctc_cell}), "
                f"{get_column_letter(col_index('Gross Salary (Monthly)'))}{row_idx}+SUM({', '.join(benefits_sum)}), \"\")"
            )
            ws.cell(row=row_idx, column=col_index('Total CTC (Annually)')).value = (
                f"=IF(ISNUMBER({get_column_letter(col_index('Total CTC (Monthly)'))}{row_idx}), "
                f"{get_column_letter(col_index('Total CTC (Monthly)'))}{row_idx}*12, \"\")"
            )

            # Net Salary (Total CTC - Deductions)
            deductions_sum = []
            for deduction in default_deductions:
                monthly_col = col_index(f'Monthly ({deduction["component_name"]})')
                deductions_sum.append(
                    f"IF(ISNUMBER({get_column_letter(monthly_col)}{row_idx}), {get_column_letter(monthly_col)}{row_idx}, 0)")
            for deduction_row in deduction_data:
                monthly_col = col_index(f'Monthly ({deduction_row["deduction_name"]})')
                deductions_sum.append(
                    f"IF(ISNUMBER({get_column_letter(monthly_col)}{row_idx}), {get_column_letter(monthly_col)}{row_idx}, 0)")

            monthly_net_cell = get_column_letter(col_index('Net Salary (Monthly)'))
            monthly_total_ctc = get_column_letter(col_index('Total CTC (Monthly)'))

            ws.cell(row=row_idx, column=col_index('Net Salary (Monthly)')).value = (
                f"=IF(ISNUMBER({annual_ctc_cell}), "
                f"MROUND({monthly_total_ctc}{row_idx}-SUM({', '.join(benefits_sum + deductions_sum)}), 1000), \"\")"
            )

            ws.cell(row=row_idx, column=col_index('Net Salary (Annually)')).value = (
                f"=IF(ISNUMBER({monthly_net_cell}{row_idx}), "
                f"MROUND({monthly_net_cell}{row_idx}*12, 1000), \"\")"
            )

        # Add data validation for tax regime
        tax_regime_col_idx = col_index('Tax Regime Opted')
        dv = DataValidation(type="list", formula1='"old,new"', allow_blank=True)
        dv.error = 'Select either "old" or "new"'
        dv.errorTitle = 'Invalid Input'
        dv.add(f"{get_column_letter(tax_regime_col_idx)}2:{get_column_letter(tax_regime_col_idx)}{n_rows}")
        ws.add_data_validation(dv)

        # Protect sheet
        ws.protection.sheet = True
        ws.protection.password = 'tara'
        ws.protection.enable()

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"EmployeeSalaryTemplate_{date.today().strftime('%Y-%m-%d')}.xlsx"
        return HttpResponse(buffer.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            headers={'Content-Disposition': f'attachment; filename="{filename}"'})

    except PayrollOrg.DoesNotExist:
        return HttpResponse("Payroll not found", status=404)






